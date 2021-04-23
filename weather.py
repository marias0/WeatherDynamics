from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from datetime import date, timedelta, datetime
import bs4
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment, NamedStyle, PatternFill
from openpyxl.chart import Reference, StockChart, LineChart

options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")

chrome_driver_path = "C:\\Program Files\\chromedriver.exe"
driver = webdriver.Chrome(executable_path=chrome_driver_path, options=options)

website = "https://www.gismeteo.ru/weather-sankt-peterburg-4079/10-days/"
driver.maximize_window()
driver.get(website)

soup = bs4.BeautifulSoup(driver.page_source, "html.parser")
spans = soup.find("div", class_="values").findAll(
    "span", class_="unit unit_temperature_c"
)

weather = []
for span in spans:
    weather.append(span.text.strip())
# print(weather)

weatherMax = []
weatherMin = []
count = 0
while count <= len(weather) - 1:
    max_weather = weather[count]
    min_weather = weather[count + 1]
    count += 2
    weatherMax.append(int(max_weather))
    weatherMin.append(int(min_weather))

today = date.today()
weatherForecast = []
for i in range(10):
    date = today + timedelta(i)
    weatherForecast.append([date, weatherMax[i], weatherMin[i]])
# print(weatherForecast)

months = []
days = []
for i in range(9):
    date = today - timedelta(i + 18)
    current_date = str(date)
    year, month, day = current_date.split("-")
    months.append(int(month))
    days.append(int(day))

months.sort(reverse=False)
months = list(set([x for x in months if months.count(x) > 1]))
print(months)

driver.quit()

weatherPast = []

for month in months:
    driver = webdriver.Chrome(executable_path=chrome_driver_path, options=options)
    website_path = "https://www.gismeteo.ru/diary/4079/2021/{query}/"
    data_website = website_path.format(query=month)
    driver.get(data_website)

    past = bs4.BeautifulSoup(driver.page_source, "html.parser")
    date_elements = past.find(id="data_block").findAll("td", class_="first")
    weather_elements = past.find(id="data_block").findAll(
        "td", class_=["first_in_group positive", "first_in_group"]
    )
    # print(weather_elements)

    dayInMonth = []
    for day in date_elements:
        dayInMonth.append(day.text.strip())

    dayAMonth = []
    for day in dayInMonth:
        Dt = datetime(int(2021), int(month), int(day))
        D = datetime.date(Dt)
        dayAMonth.append(D)
    # print(dayAMonth)

    weatherInMonth = []
    for temp in weather_elements:
        weatherInMonth.append(temp.text.strip())
    # print(len(dayInMonth))
    # print(len(weatherInMonth))

    weatherMaxM = []
    weatherMinM = []
    count = 0
    while count <= len(weatherInMonth) - 2:
        max_weather_month = weatherInMonth[count]
        min_weather_month = weatherInMonth[count + 1]
        count += 2
        weatherMaxM.append(int(max_weather_month))
        weatherMinM.append(int(min_weather_month))
    # print(weatherMaxM)
    # print(weatherMinM)

    i = 0
    for i in range(len(dayInMonth)):
        weatherPast.append([dayAMonth[i], weatherMaxM[i], weatherMinM[i]])
    # print(weatherPast)

    driver.quit()

weatherInPast = []

for k in range(9):
    date = today - timedelta(k + 1)
    current_date = str(date)
    for row in weatherPast:
        if row[0] == date:
            weatherInPast.append(row)
# print(weatherInPast)

wb = Workbook()
destination_filename = "Weather.xlsx"

ws = wb.active
ws.title = "Weather Dynamics"

ws["A1"] = "Date"
ws["B1"] = "Temperature, C"
ws["B2"] = "Max"
ws["C2"] = "Min"
ws.merge_cells("A1:A2")
ws.merge_cells("B1:C1")

for row in reversed(weatherInPast):
    ws.append(row)

for row in weatherForecast:
    ws.append(row)

ws.column_dimensions["A"].width = 11

dateStyle = NamedStyle(name="dateStyle", number_format="dd.mm.yy")
for row in ws["A3:A21"]:
    for cell in row:
        cell.style = dateStyle

cellBorder = Side(style="thin", color="808080")

for row in ws["A1:C21"]:
    for cell in row:
        cell.font = Font(name="Malgun Gothic")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=cellBorder, bottom=cellBorder, left=cellBorder, right=cellBorder
        )

today_fill = PatternFill(start_color="45F045", end_color="45F045", fill_type="solid")
ws["A12"].fill = today_fill
ws["B12"].fill = today_fill
ws["C12"].fill = today_fill

chartObject = StockChart()

labels_reference = Reference(ws, min_row=3, min_col=1, max_row=21)
values_reference = Reference(ws, min_row=3, min_col=2, max_row=21, max_col=3)

chartObject.add_data(values_reference)
chartObject.set_categories(labels_reference)

ws.add_chart(chartObject, "E3")

chartObject.title = "Weather"
chartObject.x_axis.title = "Date"
chartObject.y_axis.title = "Temperature"

chartObject.style = 5
chartObject.legend = None
chartObject.series[0].graphicalProperties.line.solidFill = "FF4500"
chartObject.series[1].graphicalProperties.line.solidFill = "1E90FF"

chartObject.height = 9
chartObject.width = 19

wb.save(destination_filename)