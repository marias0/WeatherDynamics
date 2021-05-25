from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment, NamedStyle, PatternFill
from openpyxl.chart import Reference, StockChart


# Chromedriver setup
options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")

driver = webdriver.Chrome(
    executable_path="C:\\Program Files\\chromedriver.exe", options=options
)


# Function to parse data from the website
def Data(website, find_tag, find_class, findAll_tag, findAll_class):

    driver.get(website)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    info = soup.find(find_tag, class_=find_class).findAll(
        findAll_tag, class_=findAll_class
    )

    return list(map(lambda x: x.text.strip(), info))


# Making a list with 10 days weather forecast
tenDaysWeather = Data(
    "https://www.gismeteo.ru/weather-sankt-peterburg-4079/10-days/",
    "div",
    "values",
    "span",
    "unit unit_temperature_c",
)

weatherMax = [int(i) for i in tenDaysWeather[::2]]
weatherMin = [int(i) for i in tenDaysWeather[1::2]]


TODAY = date.today()
tenDaysAhead = [TODAY + timedelta(i) for i in range(10)]
weatherForecast = list(zip(tenDaysAhead, weatherMax, weatherMin))


# Weather in past 10 days
tenDaysBack = [TODAY - timedelta(i + 1) for i in range(10)]


def get_month_weather(month):
    path = "https://www.gismeteo.ru/diary/4079/2021/{query}/"
    website_path = path.format(query=month)

    monthDays = Data(
        website_path,
        "div",
        "container",
        "td",
        "first",
    )

    monthWeather = Data(
        website_path,
        "div",
        "container",
        "td",
        ["first_in_group positive", "first_in_group"],
    )

    monthWeatherMax = monthWeather[::2]
    monthWeatherMin = monthWeather[1::2]

    return list(zip(monthDays, monthWeatherMax, monthWeatherMin))


def get_day_weather(date):
    year, month, day = str(date).split("-")
    allMonth = get_month_weather(int(month))
    for row in allMonth:
        if row[0] == str(int(day)):
            return [date, int(row[1]), int(row[2])]


# Making a list with weather in 10 days past
weatherPast = list(map(get_day_weather, tenDaysBack))


# Creating an excel file
wb = Workbook()
destination_filename = "Weather.xlsx"

ws = wb.active
ws.title = "Weather Dynamics"

ws["A1"] = "Date"
ws["B1"] = "Temperature, C"
ws["B2"] = "Max"
ws["C2"] = "Min"
ws.merge_cells("A1:A2")
ws.merge_cells("B1:C1")

for row in reversed(weatherPast):
    ws.append(row)

for row in weatherForecast:
    ws.append(row)

ws.column_dimensions["A"].width = 11

dateStyle = NamedStyle(name="dateStyle", number_format="dd.mm.yy")
for row in ws["A3:A22"]:
    for cell in row:
        cell.style = dateStyle

cellBorder = Side(style="thin", color="808080")

for row in ws["A1:C22"]:
    for cell in row:
        cell.font = Font(name="Malgun Gothic")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=cellBorder, bottom=cellBorder, left=cellBorder, right=cellBorder
        )

today_fill = PatternFill(start_color="45F045", end_color="45F045", fill_type="solid")
ws["A13"].fill = today_fill
ws["B13"].fill = today_fill
ws["C13"].fill = today_fill


chartObject = StockChart()

labels_reference = Reference(ws, min_row=3, min_col=1, max_row=22)
values_reference = Reference(ws, min_row=3, min_col=2, max_row=22, max_col=3)

chartObject.add_data(values_reference)
chartObject.set_categories(labels_reference)

ws.add_chart(chartObject, "E3")

chartObject.title = "Weather"
chartObject.x_axis.title = "Date"
chartObject.y_axis.title = "Temperature"

chartObject.style = 5
chartObject.legend = None
chartObject.series[0].graphicalProperties.line.solidFill = "FF4500"
chartObject.series[1].graphicalProperties.line.solidFill = "1E90FF"

chartObject.height = 9
chartObject.width = 19

wb.save(destination_filename)